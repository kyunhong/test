from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
import io
import openpyxl
import math
from openpyxl.styles import PatternFill, Font, Alignment
from ..database import get_db
from ..models import ExamRecord

router = APIRouter()

# ─────────────────────────────────────────
# 공통 헬퍼
# ─────────────────────────────────────────
def safe_avg(values):
    vals = [v for v in values if v is not None]
    return round(sum(vals) / len(vals), 2) if vals else None

def grade_dist(records, field):
    dist = {}
    students = {}
    for r in records:
        g = getattr(r, field)
        if g is not None:
            dist[g] = dist.get(g, 0) + 1
            if g not in students:
                students[g] = []
            students[g].append({
                "ban": r.ban, "number": r.number, "name": r.name
            })
    return dist, students

def calc_explore(society_grade, science_grade, method: str):
    """
    method:
      'avg'   → 평균 (소수점 버림)
      'best'  → 유리한 과목 (등급 숫자 작은 쪽)
      'worst' → 두 과목 모두 충족 (등급 숫자 큰 쪽)
    """
    grades = [g for g in [society_grade, science_grade] if g is not None]
    if not grades:
        return None
    if len(grades) == 1:
        return grades[0]
    if method == 'avg':
        return math.floor((grades[0] + grades[1]) / 2)
    elif method == 'best':
        return min(grades)
    elif method == 'worst':
        return max(grades)
    return None


def calc_explore_all(record):
    """레코드 하나에 대해 세 방식 모두 계산"""
    s = record.society_grade
    c = record.science_grade
    return {
        'avg':   calc_explore(s, c, 'avg'),
        'best':  calc_explore(s, c, 'best'),
        'worst': calc_explore(s, c, 'worst'),
    }

# ─────────────────────────────────────────
# 1. 전체 학생 목록
# ─────────────────────────────────────────
@router.get("/analysis/all-students")
def get_all_students(exam_name: str, db: Session = Depends(get_db)):
    records = db.query(ExamRecord).filter(
        ExamRecord.exam_name == exam_name
    ).order_by(ExamRecord.ban, ExamRecord.number).all()

    return [{
        "ban": r.ban, "number": r.number, "name": r.name,
        "korean_std": r.korean_std, "korean_pct": r.korean_pct,
        "korean_grade": r.korean_grade,
        "math_std": r.math_std, "math_pct": r.math_pct,
        "math_grade": r.math_grade,
        "english_grade": r.english_grade,
        "society_std": r.society_std, "society_pct": r.society_pct,
        "society_grade": r.society_grade,
        "science_std": r.science_std, "science_pct": r.science_pct,
        "science_grade": r.science_grade,
        "history_grade": r.history_grade,
    } for r in records]


# ─────────────────────────────────────────
# 2. 성적 분석 요약
# ─────────────────────────────────────────
@router.get("/analysis/summary")
def get_summary(exam_name: str, db: Session = Depends(get_db)):
    records = db.query(ExamRecord).filter(
        ExamRecord.exam_name == exam_name
    ).all()

    if not records:
        return {}

    def subject_summary(std_field, pct_field, grade_field):
        stds   = [getattr(r, std_field)   for r in records if getattr(r, std_field)   is not None]
        pcts   = [getattr(r, pct_field)   for r in records if getattr(r, pct_field)   is not None]
        dist, students = grade_dist(records, grade_field)
        return {
            "avg_std":      round(sum(stds)/len(stds), 2) if stds else None,
            "avg_pct":      round(sum(pcts)/len(pcts), 2) if pcts else None,
            "grade1_count": dist.get(1, 0),
            "grade_dist":   dist,
            "grade_students": students,
        }

    def grade_only_summary(grade_field):
        dist, students = grade_dist(records, grade_field)
        return {
            "avg_std": None, "avg_pct": None,
            "grade1_count": dist.get(1, 0),
            "grade_dist": dist,
            "grade_students": students,
        }

    return {
        "total":   len(records),
        "korean":  subject_summary("korean_std",  "korean_pct",  "korean_grade"),
        "math":    subject_summary("math_std",    "math_pct",    "math_grade"),
        "english": grade_only_summary("english_grade"),
        "society": subject_summary("society_std", "society_pct", "society_grade"),
        "science": subject_summary("science_std", "science_pct", "science_grade"),
        "history": grade_only_summary("history_grade"),
    }


# ─────────────────────────────────────────
# 3. 석차
# ─────────────────────────────────────────
@router.get("/analysis/ranking")
def get_ranking(
    exam_name: str,
    mode: str = Query("std"),
    subjects: str = Query("korean,math,english,society,science,history"),
    db: Session = Depends(get_db)
):
    records = db.query(ExamRecord).filter(
        ExamRecord.exam_name == exam_name
    ).all()

    selected = [s.strip() for s in subjects.split(",") if s.strip()]

    std_map = {
        "korean": "korean_std", "math": "math_std",
        "society": "society_std", "science": "science_std",
    }
    grade_map = {
        "korean": "korean_grade", "math": "math_grade",
        "english": "english_grade", "society": "society_grade",
        "science": "science_grade", "history": "history_grade",
    }

    result = []
    for r in records:
        if mode == "std":
            vals = [
                getattr(r, std_map[s])
                for s in selected if s in std_map
                and getattr(r, std_map[s]) is not None
            ]
            sort_val = round(sum(vals), 2) if vals else 0
        else:
            vals = [
                getattr(r, grade_map[s])
                for s in selected if s in grade_map
                and getattr(r, grade_map[s]) is not None
            ]
            sort_val = round(sum(vals) / len(vals), 2) if vals else 9

        result.append({
            "ban": r.ban, "number": r.number, "name": r.name,
            "korean_std": r.korean_std, "korean_grade": r.korean_grade,
            "math_std": r.math_std, "math_grade": r.math_grade,
            "english_grade": r.english_grade,
            "society_std": r.society_std, "society_grade": r.society_grade,
            "science_std": r.science_std, "science_grade": r.science_grade,
            "history_grade": r.history_grade,
            "sort_val": sort_val,
        })

    if mode == "std":
        result.sort(key=lambda x: x["sort_val"], reverse=True)
    else:
        result.sort(key=lambda x: x["sort_val"])

    for i, r in enumerate(result):
        r["rank"] = i + 1

    return result


# ─────────────────────────────────────────
# 4. 대시보드 요약
# ─────────────────────────────────────────
@router.get("/analysis/dashboard")
def get_dashboard(exam_name: str, db: Session = Depends(get_db)):
    records = db.query(ExamRecord).filter(
        ExamRecord.exam_name == exam_name
    ).all()

    if not records:
        return {}

    total = len(records)

    def g1_count(field):
        return sum(1 for r in records if getattr(r, field) == 1)

    def avg_std(field):
        vals = [getattr(r, field) for r in records if getattr(r, field) is not None]
        return round(sum(vals)/len(vals), 2) if vals else None

    def avg_grade(field):
        vals = [getattr(r, field) for r in records if getattr(r, field) is not None]
        return round(sum(vals)/len(vals), 2) if vals else None

    # 반별 인원
    ban_counts = {}
    for r in records:
        ban_counts[r.ban] = ban_counts.get(r.ban, 0) + 1

    # 전 과목 1등급 학생
    all_grade1 = [
        {"ban": r.ban, "number": r.number, "name": r.name}
        for r in records
        if all([
            r.korean_grade == 1, r.math_grade == 1,
            r.english_grade == 1, r.society_grade == 1,
            r.science_grade == 1
        ])
    ]

    return {
        "total": total,
        "ban_count": len(ban_counts),
        "subjects": {
            "korean":  {"g1": g1_count("korean_grade"),  "avg_std": avg_std("korean_std"),  "avg_grade": avg_grade("korean_grade")},
            "math":    {"g1": g1_count("math_grade"),    "avg_std": avg_std("math_std"),    "avg_grade": avg_grade("math_grade")},
            "english": {"g1": g1_count("english_grade"), "avg_std": None,                   "avg_grade": avg_grade("english_grade")},
            "society": {"g1": g1_count("society_grade"), "avg_std": avg_std("society_std"), "avg_grade": avg_grade("society_grade")},
            "science": {"g1": g1_count("science_grade"), "avg_std": avg_std("science_std"), "avg_grade": avg_grade("science_grade")},
            "history": {"g1": g1_count("history_grade"), "avg_std": None,                   "avg_grade": avg_grade("history_grade")},
        },
        "all_grade1": all_grade1,
    }


# ─────────────────────────────────────────
# 5. 반별 비교
# ─────────────────────────────────────────
@router.get("/analysis/class-compare")
def get_class_compare(exam_name: str, db: Session = Depends(get_db)):
    records = db.query(ExamRecord).filter(
        ExamRecord.exam_name == exam_name
    ).all()

    if not records:
        return []

    bans = sorted(set(r.ban for r in records))

    result = []
    for ban in bans:
        ban_records = [r for r in records if r.ban == ban]
        total = len(ban_records)

        def g1(field):
            return sum(1 for r in ban_records if getattr(r, field) == 1)

        def avg_s(field):
            vals = [getattr(r, field) for r in ban_records if getattr(r, field) is not None]
            return round(sum(vals)/len(vals), 2) if vals else None

        def avg_g(field):
            vals = [getattr(r, field) for r in ban_records if getattr(r, field) is not None]
            return round(sum(vals)/len(vals), 2) if vals else None

        result.append({
            "ban": ban,
            "total": total,
            "korean":  {"avg_std": avg_s("korean_std"),  "avg_grade": avg_g("korean_grade"),  "g1": g1("korean_grade")},
            "math":    {"avg_std": avg_s("math_std"),    "avg_grade": avg_g("math_grade"),    "g1": g1("math_grade")},
            "english": {"avg_std": None,                  "avg_grade": avg_g("english_grade"), "g1": g1("english_grade")},
            "society": {"avg_std": avg_s("society_std"), "avg_grade": avg_g("society_grade"), "g1": g1("society_grade")},
            "science": {"avg_std": avg_s("science_std"), "avg_grade": avg_g("science_grade"), "g1": g1("science_grade")},
            "history": {"avg_std": None,                  "avg_grade": avg_g("history_grade"), "g1": g1("history_grade")},
        })

    return result


# ─────────────────────────────────────────
# 6. 특이 학생 감지
# ─────────────────────────────────────────
# ── 특이 학생 감지 ────────────────────────────────────
@router.get("/analysis/alert-students")
def get_alert_students(
    exam_name: str,
    threshold: int = 4,
    explore_method: str = 'avg',   # avg | best | worst
    db: Session = Depends(get_db)
):
    records = db.query(ExamRecord).filter(
        ExamRecord.exam_name == exam_name
    ).all()

    alerts = []
    for r in records:
        explore = calc_explore(r.society_grade, r.science_grade, explore_method)
        explore_all = calc_explore_all(r)

        # 과목별 등급 딕셔너리
        subject_grades = {
            '국어':     r.korean_grade,
            '수학':     r.math_grade,
            '영어':     r.english_grade,
            '탐구':     explore,
            '한국사':   r.history_grade,
        }

        # threshold 이하인 과목 목록
        weak_subjects = [
            subj for subj, grade in subject_grades.items()
            if grade is not None and grade >= threshold
        ]

        if weak_subjects:
            alerts.append({
                'ban':            r.ban,
                'number':         r.number,
                'name':           r.name,
                'korean_grade':   r.korean_grade,
                'math_grade':     r.math_grade,
                'english_grade':  r.english_grade,
                'society_grade':  r.society_grade,
                'science_grade':  r.science_grade,
                'history_grade':  r.history_grade,
                # 탐구 합산값 세 가지 모두 내려줌
                'explore_avg':    explore_all['avg'],
                'explore_best':   explore_all['best'],
                'explore_worst':  explore_all['worst'],
                # 현재 선택된 방식의 탐구 등급
                'explore_grade':  explore,
                'weak_subjects':  weak_subjects,
                'weak_count':     len(weak_subjects),
            })

    # 위험 과목 많은 순 → 반 → 번호 순 정렬
    alerts.sort(key=lambda x: (-x['weak_count'], x['ban'], x['number']))
    return alerts


# ─────────────────────────────────────────
# 7. 시험 간 상세 비교 (수정)
# ─────────────────────────────────────────
@router.get("/analysis/detail-compare")
def get_detail_compare(
    exam_name: str,
    prev_exam_name: str,
    db: Session = Depends(get_db)
):
    curr_records = db.query(ExamRecord).filter(
        ExamRecord.exam_name == exam_name
    ).all()
    prev_records = db.query(ExamRecord).filter(
        ExamRecord.exam_name == prev_exam_name
    ).all()

    prev_map = {r.name: r for r in prev_records}

    grade_fields = [
        "korean_grade", "math_grade", "english_grade",
        "society_grade", "science_grade", "history_grade"
    ]
    std_fields = ["korean_std", "math_std", "society_std", "science_std"]

    # ── 학생별 비교 ──
    students = []
    for r in curr_records:
        prev = prev_map.get(r.name)

        def diff_grade(field):
            c = getattr(r, field)
            p = getattr(prev, field) if prev else None
            if c is None or p is None:
                return None
            return p - c

        def diff_std(field):
            c = getattr(r, field)
            p = getattr(prev, field) if prev else None
            if c is None or p is None:
                return None
            return round(c - p, 2)

        grade_diffs  = [diff_grade(f) for f in grade_fields]
        valid_diffs  = [d for d in grade_diffs if d is not None]
        total_improvement = sum(valid_diffs) if valid_diffs else 0

        students.append({
            "ban": r.ban, "number": r.number, "name": r.name,
            "korean_grade":  r.korean_grade,  "korean_std":  r.korean_std,
            "math_grade":    r.math_grade,    "math_std":    r.math_std,
            "english_grade": r.english_grade,
            "society_grade": r.society_grade, "society_std": r.society_std,
            "science_grade": r.science_grade, "science_std": r.science_std,
            "history_grade": r.history_grade,
            "diff_korean_grade":  diff_grade("korean_grade"),
            "diff_math_grade":    diff_grade("math_grade"),
            "diff_english_grade": diff_grade("english_grade"),
            "diff_society_grade": diff_grade("society_grade"),
            "diff_science_grade": diff_grade("science_grade"),
            "diff_history_grade": diff_grade("history_grade"),
            "diff_korean_std":  diff_std("korean_std"),
            "diff_math_std":    diff_std("math_std"),
            "diff_society_std": diff_std("society_std"),
            "diff_science_std": diff_std("science_std"),
            "total_improvement": total_improvement,
            "has_prev": prev is not None,
        })

    students.sort(key=lambda x: x["total_improvement"], reverse=True)
    for i, s in enumerate(students):
        s["improve_rank"] = i + 1

    # ── 과목별 평균 비교 ──
    def subj_avg(records_list, field):
        vals = [getattr(r, field) for r in records_list if getattr(r, field) is not None]
        return round(sum(vals) / len(vals), 2) if vals else None

    subject_compare = {
        "korean":  {
            "curr_std":   subj_avg(curr_records, "korean_std"),
            "prev_std":   subj_avg(prev_records, "korean_std"),
            "curr_grade": subj_avg(curr_records, "korean_grade"),
            "prev_grade": subj_avg(prev_records, "korean_grade"),
        },
        "math": {
            "curr_std":   subj_avg(curr_records, "math_std"),
            "prev_std":   subj_avg(prev_records, "math_std"),
            "curr_grade": subj_avg(curr_records, "math_grade"),
            "prev_grade": subj_avg(prev_records, "math_grade"),
        },
        "english": {
            "curr_std": None, "prev_std": None,
            "curr_grade": subj_avg(curr_records, "english_grade"),
            "prev_grade": subj_avg(prev_records, "english_grade"),
        },
        "society": {
            "curr_std":   subj_avg(curr_records, "society_std"),
            "prev_std":   subj_avg(prev_records, "society_std"),
            "curr_grade": subj_avg(curr_records, "society_grade"),
            "prev_grade": subj_avg(prev_records, "society_grade"),
        },
        "science": {
            "curr_std":   subj_avg(curr_records, "science_std"),
            "prev_std":   subj_avg(prev_records, "science_std"),
            "curr_grade": subj_avg(curr_records, "science_grade"),
            "prev_grade": subj_avg(prev_records, "science_grade"),
        },
        "history": {
            "curr_std": None, "prev_std": None,
            "curr_grade": subj_avg(curr_records, "history_grade"),
            "prev_grade": subj_avg(prev_records, "history_grade"),
        },
    }

    # ── 과목별 등급 분포 (1~9등급 인원수) ──
    subject_keys = {
        "korean":  "korean_grade",
        "math":    "math_grade",
        "english": "english_grade",
        "society": "society_grade",
        "science": "science_grade",
        "history": "history_grade",
    }

    def grade_distribution(records_list, field):
        dist = {g: 0 for g in range(1, 10)}
        for r in records_list:
            g = getattr(r, field)
            if g is not None and 1 <= g <= 9:
                dist[g] += 1
        return dist

    def grade_band_count(records_list, field):
        """1~2등급 / 3~4등급 / 5등급이하 구간별 인원"""
        top    = sum(1 for r in records_list if getattr(r, field) in [1, 2])
        mid    = sum(1 for r in records_list if getattr(r, field) in [3, 4])
        low    = sum(1 for r in records_list if getattr(r, field) is not None and getattr(r, field) >= 5)
        return {"top": top, "mid": mid, "low": low}

    grade_dist_compare = {}
    for subj, field in subject_keys.items():
        grade_dist_compare[subj] = {
            "curr_dist":  grade_distribution(curr_records, field),
            "prev_dist":  grade_distribution(prev_records, field),
            "curr_band":  grade_band_count(curr_records, field),
            "prev_band":  grade_band_count(prev_records, field),
            # 1등급 인원만 따로
            "curr_g1": sum(1 for r in curr_records if getattr(r, field) == 1),
            "prev_g1": sum(1 for r in prev_records if getattr(r, field) == 1),
        }

    # ── 학교 전체 요약 ──
    def total_g1(records_list):
        count = 0
        for f in subject_keys.values():
            count += sum(1 for r in records_list if getattr(r, f) == 1)
        return count

    def avg_all_grades(records_list):
        vals = []
        for f in subject_keys.values():
            vals += [getattr(r, f) for r in records_list if getattr(r, f) is not None]
        return round(sum(vals) / len(vals), 2) if vals else None

    school_summary = {
        "curr_total":      len(curr_records),
        "prev_total":      len(prev_records),
        "curr_total_g1":   total_g1(curr_records),
        "prev_total_g1":   total_g1(prev_records),
        "curr_avg_grade":  avg_all_grades(curr_records),
        "prev_avg_grade":  avg_all_grades(prev_records),
    }

    return {
        "students":          students,
        "subject_compare":   subject_compare,
        "grade_dist_compare": grade_dist_compare,
        "school_summary":    school_summary,
        "curr_total":        len(curr_records),
        "prev_total":        len(prev_records),
    }


# ─────────────────────────────────────────
# 8. 엑셀 내보내기
# ─────────────────────────────────────────
@router.get("/analysis/export-excel")
def export_excel(exam_name: str, db: Session = Depends(get_db)):
    records = db.query(ExamRecord).filter(
        ExamRecord.exam_name == exam_name
    ).order_by(ExamRecord.ban, ExamRecord.number).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = exam_name[:30]

    headers = [
        "반", "번호", "이름",
        "국어표준", "국어백분위", "국어등급",
        "수학표준", "수학백분위", "수학등급",
        "영어등급",
        "통합사회표준", "통합사회백분위", "통합사회등급",
        "통합과학표준", "통합과학백분위", "통합과학등급",
        "한국사등급"
    ]

    # 헤더 스타일
    header_fill = PatternFill("solid", fgColor="BFD7FF")
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 등급 색상
    grade_colors = {
        1: "FEE2E2", 2: "FEF3C7", 3: "ECFDF5",
        7: "F3F4F6", 8: "F3F4F6", 9: "F3F4F6"
    }

    for row_idx, r in enumerate(records, 2):
        row_data = [
            r.ban, r.number, r.name,
            r.korean_std, r.korean_pct, r.korean_grade,
            r.math_std, r.math_pct, r.math_grade,
            r.english_grade,
            r.society_std, r.society_pct, r.society_grade,
            r.science_std, r.science_pct, r.science_grade,
            r.history_grade,
        ]
        grade_cols = [6, 9, 10, 13, 16, 17]  # 등급 열 인덱스

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")
            if col_idx in grade_cols and val in grade_colors:
                cell.fill = PatternFill("solid", fgColor=grade_colors[val])

    # 열 너비 자동 조정
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"{exam_name}_성적.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )